import os
import duckdb
from typing import Literal, Optional
import io
from contextlib import redirect_stdout
from tabulate import tabulate
import logging
import time
import json
from .configs import SERVER_VERSION

logger = logging.getLogger("mcp_server_motherduck")


class DatabaseClient:
    def __init__(
        self,
        db_path: str | None = None,
        motherduck_token: str | None = None,
        home_dir: str | None = None,
        saas_mode: bool = False,
        read_only: bool = False,
    ):
        self._read_only = read_only
        self.db_path, self.db_type = self._resolve_db_path_type(
            db_path, motherduck_token, saas_mode
        )
        logger.info(f"Database client initialized in `{self.db_type}` mode")

        # Set the home directory for DuckDB
        if home_dir:
            os.environ["HOME"] = home_dir

        self.conn = self._initialize_connection()

    def _initialize_connection(self) -> Optional[duckdb.DuckDBPyConnection]:
        """Initialize connection to the MotherDuck or DuckDB database"""

        logger.info(f"🔌 Connecting to {self.db_type} database")

        # S3 databases don't support read-only mode
        if self.db_type == "s3" and self._read_only:
            raise ValueError("Read-only mode is not supported for S3 databases")

        if self.db_type == "duckdb" and self._read_only:
            # check that we can connect, issue a `select 1` and then close + return None
            try:
                conn = duckdb.connect(
                    self.db_path,
                    config={
                        "custom_user_agent": f"mcp-server-motherduck/{SERVER_VERSION}"
                    },
                    read_only=self._read_only,
                )
                conn.execute("SELECT 1")
                conn.close()
                return None
            except Exception as e:
                logger.error(f"❌ Read-only check failed: {e}")
                raise

        # Check if this is an S3 path
        if self.db_type == "s3":
            # For S3, we need to create an in-memory connection and attach the S3 database
            conn = duckdb.connect(':memory:')
            
            # Install and load the httpfs extension for S3 support
            import io
            from contextlib import redirect_stdout, redirect_stderr
            
            null_file = io.StringIO()
            with redirect_stdout(null_file), redirect_stderr(null_file):
                try:
                    conn.execute("INSTALL httpfs;")
                except:
                    pass  # Extension might already be installed
                conn.execute("LOAD httpfs;")
            
            # Install and load the excel extension for reading .xlsx files
            with redirect_stdout(null_file), redirect_stderr(null_file):
                try:
                    conn.execute("INSTALL excel;")
                except:
                    pass  # Extension might already be installed
                conn.execute("LOAD excel;")
            
            # Configure S3 credentials from environment variables using CREATE SECRET
            aws_access_key = os.environ.get('AWS_ACCESS_KEY_ID')
            aws_secret_key = os.environ.get('AWS_SECRET_ACCESS_KEY')
            aws_region = os.environ.get('AWS_DEFAULT_REGION', 'us-east-1')
            
            
            if aws_access_key and aws_secret_key:
                # Use CREATE SECRET for better credential management
                conn.execute(f"""
                    CREATE SECRET IF NOT EXISTS s3_secret (
                        TYPE S3,
                        KEY_ID '{aws_access_key}',
                        SECRET '{aws_secret_key}',
                        REGION '{aws_region}'
                    );
                """)
            
            # Attach the S3 database
            try:
                # For S3, we always attach as READ_ONLY since S3 storage is typically read-only
                # Even when not in read_only mode, we attach as READ_ONLY for S3
                conn.execute(f"ATTACH '{self.db_path}' AS s3db (READ_ONLY);")
                # Use the attached database
                conn.execute("USE s3db;")
                logger.info(f"✅ Successfully connected to {self.db_type} database (attached as read-only)")
            except Exception as e:
                logger.error(f"Failed to attach S3 database: {e}")
                # If the database doesn't exist and we're not in read-only mode, try to create it
                if "database does not exist" in str(e) and not self._read_only:
                    logger.info("S3 database doesn't exist, attempting to create it...")
                    try:
                        # Create a new database at the S3 location
                        conn.execute(f"ATTACH '{self.db_path}' AS s3db;")
                        conn.execute("USE s3db;")
                        logger.info(f"✅ Created new S3 database at {self.db_path}")
                    except Exception as create_error:
                        logger.error(f"Failed to create S3 database: {create_error}")
                        raise
                else:
                    raise
                
            return conn

        conn = duckdb.connect(
            self.db_path,
            config={"custom_user_agent": f"mcp-server-motherduck/{SERVER_VERSION}"},
            read_only=self._read_only,
        )
        
        # Install and load the excel extension for reading .xlsx files
        import io
        from contextlib import redirect_stdout, redirect_stderr
        
        null_file = io.StringIO()
        with redirect_stdout(null_file), redirect_stderr(null_file):
            try:
                conn.execute("INSTALL excel;")
            except:
                pass  # Extension might already be installed
            conn.execute("LOAD excel;")

        logger.info(f"✅ Successfully connected to {self.db_type} database")

        return conn

    def _resolve_db_path_type(
        self, db_path: str, motherduck_token: str | None = None, saas_mode: bool = False
    ) -> tuple[str, Literal["duckdb", "motherduck", "s3"]]:
        """Resolve and validate the database path"""
        # Handle S3 paths
        if db_path.startswith("s3://"):
            return db_path, "s3"
        
        # Handle MotherDuck paths
        if db_path.startswith("md:"):
            if motherduck_token:
                logger.info("Using MotherDuck token to connect to database `md:`")
                if saas_mode:
                    logger.info("Connecting to MotherDuck in SaaS mode")
                    return (
                        f"{db_path}?motherduck_token={motherduck_token}&saas_mode=true",
                        "motherduck",
                    )
                else:
                    return (
                        f"{db_path}?motherduck_token={motherduck_token}",
                        "motherduck",
                    )
            elif os.getenv("motherduck_token"):
                logger.info(
                    "Using MotherDuck token from env to connect to database `md:`"
                )
                return (
                    f"{db_path}?motherduck_token={os.getenv('motherduck_token')}",
                    "motherduck",
                )
            else:
                raise ValueError(
                    "Please set the `motherduck_token` as an environment variable or pass it as an argument with `--motherduck-token` when using `md:` as db_path."
                )

        if db_path == ":memory:":
            return db_path, "duckdb"

        return db_path, "duckdb"

    def _execute(self, query: str) -> str:
        if self.conn is None:
            # open short lived readonly connection for local DuckDB, run query, close connection, return result
            conn = duckdb.connect(
                self.db_path,
                config={"custom_user_agent": f"mcp-server-motherduck/{SERVER_VERSION}"},
                read_only=self._read_only,
            )
            q = conn.execute(query)
        else:
            q = self.conn.execute(query)

        out = tabulate(
            q.fetchall(),
            headers=[d[0] + "\n" + str(d[1]) for d in q.description],
            tablefmt="pretty",
        )

        if self.conn is None:
            conn.close()

        return out

    def _execute_json(self, query: str) -> dict:
        """Execute query and return structured JSON response"""
        start_time = time.time()
        
        if self.conn is None:
            # open short lived readonly connection for local DuckDB, run query, close connection, return result
            conn = duckdb.connect(
                self.db_path,
                config={"custom_user_agent": f"mcp-server-motherduck/{SERVER_VERSION}"},
                read_only=self._read_only,
            )
            q = conn.execute(query)
        else:
            q = self.conn.execute(query)

        # Fetch results as DataFrame
        df = q.fetchdf()
        
        # Limit to 1000 rows to prevent context overflow
        truncated = False
        if len(df) > 1000:
            df = df.head(1000)
            truncated = True

        execution_time = int((time.time() - start_time) * 1000)  # milliseconds

        if self.conn is None:
            conn.close()

        return {
            "success": True,
            "data": df.to_dict(orient="records"),
            "columns": list(df.columns),
            "rowCount": len(df),
            "executionTime": execution_time,
            "truncated": truncated,
            "query": query
        }

    def query(self, query: str) -> str:
        try:
            return self._execute(query)

        except Exception as e:
            raise ValueError(f"❌ Error executing query: {e}")

    def query_json(self, query: str) -> dict:
        """Execute query and return JSON response"""
        try:
            return self._execute_json(query)
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "query": query,
                "data": [],
                "columns": [],
                "rowCount": 0,
                "executionTime": 0,
                "truncated": False
            }

    def discover_excel_structure(self, file_path: str, sheet_filter: str = "*", sample_rows: int = 5) -> dict:
        """Discover structure of Excel sheets with schema and sample data"""
        try:
            import openpyxl
            
            # Verificar se arquivo existe
            if not os.path.exists(file_path):
                return {
                    "success": False,
                    "error": f"File not found: {file_path}",
                    "sheets": {}
                }
            
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # Determinar quais sheets analisar
            if sheet_filter == "*":
                target_sheets = wb.sheetnames
            else:
                target_sheets = [sheet_filter] if sheet_filter in wb.sheetnames else []
            
            if not target_sheets:
                wb.close()
                return {
                    "success": False,
                    "error": f"Sheet '{sheet_filter}' not found",
                    "availableSheets": wb.sheetnames,
                    "sheets": {}
                }
            
            sheets_data = {}
            
            for sheet_name in target_sheets:
                try:
                    # Usar DuckDB para análise eficiente
                    query = f"""
                    SELECT * 
                    FROM read_xlsx('{file_path}', 
                                  sheet='{sheet_name}', 
                                  all_varchar=true, 
                                  ignore_errors=true) 
                    LIMIT {sample_rows}
                    """
                    
                    sample_result = self._execute_json(query)
                    
                    if not sample_result.get("success"):
                        sheets_data[sheet_name] = {
                            "error": sample_result.get("error"),
                            "columns": [],
                            "rowCount": 0,
                            "sampleData": []
                        }
                        continue
                    
                    sample_data = sample_result.get("data", [])
                    columns = sample_result.get("columns", [])
                    
                    # Análise de colunas
                    columns_info = []
                    for col in columns:
                        col_values = [row.get(col) for row in sample_data if row.get(col) is not None]
                        
                        columns_info.append({
                            "name": col,
                            "type": self._infer_column_type(col_values),
                            "distinctCount": len(set(str(v) for v in col_values)),
                            "nonNullCount": len(col_values)
                        })
                    
                    # Obter total de linhas (sem limite)
                    count_query = f"""
                    SELECT COUNT(*) as total
                    FROM read_xlsx('{file_path}', 
                                  sheet='{sheet_name}', 
                                  all_varchar=true, 
                                  ignore_errors=true)
                    """
                    count_result = self._execute_json(count_query)
                    total_rows = count_result.get("data", [{}])[0].get("total", 0) if count_result.get("success") else 0
                    
                    sheets_data[sheet_name] = {
                        "columns": columns_info,
                        "rowCount": total_rows,
                        "sampleData": sample_data
                    }
                    
                except Exception as e:
                    logger.error(f"Error analyzing sheet {sheet_name}: {e}")
                    sheets_data[sheet_name] = {
                        "error": str(e),
                        "columns": [],
                        "rowCount": 0,
                        "sampleData": []
                    }
            
            wb.close()
            
            return {
                "success": True,
                "fileId": os.path.basename(file_path).replace(".xlsx", ""),
                "sheets": sheets_data
            }
            
        except Exception as e:
            logger.error(f"Error discovering Excel structure: {e}")
            return {
                "success": False,
                "error": str(e),
                "sheets": {}
            }

    def _infer_column_type(self, values: list) -> str:
        """Infer data type from sample values"""
        if not values:
            return "VARCHAR"
        
        # Tentar inferir tipo baseado nos valores
        try:
            # Verificar se todos são números
            numeric_count = sum(1 for v in values if str(v).replace('.', '', 1).replace('-', '', 1).isdigit())
            if numeric_count == len(values):
                # Verificar se tem pontos decimais
                if any('.' in str(v) for v in values):
                    return "DOUBLE"
                return "INTEGER"
        except:
            pass
        
        return "VARCHAR"
